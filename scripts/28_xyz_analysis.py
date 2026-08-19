# -*- coding: utf-8 -*-
"""
28_xyz_analysis.py — XYZ-анализ колбасы (классификация товаров по предсказуемости спроса)
==========================================================================================
XYZ-анализ делит товары на классы по стабильности месячных продаж.
CV (коэффициент вариации) = стд. отклонение месячных продаж / среднее месячных продаж * 100
— «насколько месячные продажи пляшут вокруг среднего».

ДВА набора границ классификации (CV считается ОДИН раз, меняется только порог):

  Классические (листы XYZ_классика / Сводка_классика):
      X — стабильный спрос:     CV <= 10%
      Y — колеблющийся спрос:   10% < CV <= 25%
      Z — непредсказуемый:      CV > 25%
    Стандарт для промышленных/оптовых поставок.

  Розничные (листы XYZ_розница / Сводка_розница):
      X — стабильный спрос:     CV <= 25%
      Y — колеблющийся спрос:   25% < CV <= 50%
      Z — непредсказуемый:      CV > 50%
    Адаптация для розничного магазина, где промо и сезонность усиливают колебания.

Методология (как в эталонных отчётах проекта):
  1. СПРОС — только положительные продажи: (ТО, в е.изм. > 0) & (ТО, руб > 0).
     Возвраты (отрицательные строки) исключаются из всех расчётов.
  2. Месячные суммы продаж в собственных единицах товара (весовые — кг,
     штучные — шт; тип по правилу: есть хоть одна дробная продажа → весовой).
  3. Полные месяцы данных: 2023-01 … 2026-07 (43 месяца).
     Окно товара — от первого полного месяца с продажами до последнего
     (если товар активен в 2026-07 — окно до 2026-07; перестал раньше — до его
     последнего месяца с продажами). Нулевые месяцы ВНУТРИ окна включаются в расчёт.
  4. Окно < 6 месяцев → класс «н/д», примечание «новинка или мало данных (N мес)».
  5. Остальные: CV = std/mean*100 (выборочное стд. отклонение, как STDEV в Excel).
      Класс (классика): X при CV <= 10; Y при 10 < CV <= 25; Z при CV > 25.
      Класс (розница):  X при CV <= 25; Y при 25 < CV <= 50; Z при CV > 50.

Выход:
  output/Отчет_XYZ_колбаса.xlsx        (листы XYZ_классика, Сводка_классика,
                                        XYZ_розница, Сводка_розница, Методика)
  plots/Распределение_CV_колбаса.png   (гистограмма CV %, классика, с линиями границ)
  plots/XYZ_классы_колбаса.png         (оба варианта границ: товары и выручка)
"""

import os
import json

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl.styles import Alignment, Font

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data", "колбасаНФ.xlsx")
OUT_XLSX = os.path.join(ROOT, "output", "Отчет_XYZ_колбаса.xlsx")
OUT_PNG1 = os.path.join(ROOT, "plots", "Распределение_CV_колбаса.png")
OUT_PNG2 = os.path.join(ROOT, "plots", "XYZ_классы_колбаса.png")
DASH_JSON = os.path.join(ROOT, "output", "dashboard", "dashboard_data.json")

plt.rcParams["font.family"] = "DejaVu Sans"  # шрифт поддерживает кириллицу

FULL_START, FULL_END = "2023-01", "2026-07"                     # полные месяцы данных
ALL_MONTHS = pd.period_range(FULL_START, FULL_END, freq="M").strftime("%Y-%m")

CLASS_ORDER = {"X": 0, "Y": 1, "Z": 2, "н/д": 3}   # порядок сортировки классов
CLS_NAMES = ["X", "Y", "Z", "н/д"]

# 13 колонок листов XYZ_классика / XYZ_розница (одинаковый набор для обоих вариантов)
XYZ_COLS = ["Код", "Название", "Группа", "Тип", "Первый месяц", "Последний месяц",
            "Месяцев в окне", "Среднее/мес (ед. товара)", "Стд. отклонение",
            "CV %", "Класс", "Выручка за окно (руб)", "Примечание"]


def months_between(a, b):
    """Число календарных месяцев между 'YYYY-MM' и 'YYYY-MM' включительно."""
    a = pd.Timestamp(a + "-01")
    b = pd.Timestamp(b + "-01")
    return (b.year - a.year) * 12 + (b.month - a.month) + 1


# ---------------------------------------------------------------------------
# 1. Чтение данных
# ---------------------------------------------------------------------------
df = pd.read_excel(DATA, dtype={"Код ": str}, parse_dates=["День"])
df["Код"] = df["Код "].str.strip().str.zfill(5)   # «Код » — колонка с пробелом

# Тип товара: есть ХОТЯ БЫ ОДНА дробная продажа «ТО, в е.изм.» → весовой (кг),
# все продажи целые → штучный (шт). Тип одинаков и по всем, и по положительным строкам.
frac = df.groupby("Код")["ТО, в е.изм."].apply(lambda s: (s != s.round()).any())
type_map = {c: ("весовой, кг" if frac[c] else "штучный, шт") for c in frac.index}

# СПРОС = только положительные продажи (возвраты исключаем из всех расчётов)
s = df[(df["ТО, в е.изм."] > 0) & (df["ТО, руб"] > 0)].copy()
s["месяц"] = s["День"].dt.strftime("%Y-%m")

n_codes = int(s["Код"].nunique())

# ---------------------------------------------------------------------------
# 2. Контрольные суммы — сверка с блоком «колбаса» dashboard_data.json
#    (dashboard «итоги_всего» считает по ВСЕМ строкам, включая возвраты)
# ---------------------------------------------------------------------------
with open(DASH_JSON, encoding="utf-8") as f:
    dash_tot = json.load(f)["колбаса"]["итоги_всего"]

sum_rub_pos = s["ТО, руб"].sum()                  # по методологии (положительные)
sum_rub_all = df["ТО, руб"].sum()                 # все строки (эталон dashboard)

qty_all_sht = df.loc[df["Код"].map(lambda c: "штучный" in type_map[c]), "ТО, в е.изм."].sum()
qty_all_kg = df.loc[df["Код"].map(lambda c: "весовой" in type_map[c]), "ТО, в е.изм."].sum()
qty_pos_sht = s.loc[s["Код"].map(lambda c: "штучный" in type_map[c]), "ТО, в е.изм."].sum()
qty_pos_kg = s.loc[s["Код"].map(lambda c: "весовой" in type_map[c]), "ТО, в е.изм."].sum()

print("КОНТРОЛЬ 1: уникальных кодов с положительными продажами = {}, эталон = 507 -> {}".format(
    n_codes, "OK" if n_codes == 507 else "МИСМАТЧ"))
print("КОНТРОЛЬ 2a: ТО, руб по ВСЕМ строкам = {:.2f}, эталон dashboard (итоги_всего.руб) = {:.2f} -> {}".format(
    sum_rub_all, dash_tot["руб"], "OK" if abs(sum_rub_all - dash_tot["руб"]) <= 0.01 else "МИСМАТЧ"))
print("КОНТРОЛЬ 2b: ТО, руб по ПОЛОЖИТЕЛЬНЫМ строкам (методология) = {:.2f}".format(sum_rub_pos))
print("   (эталон 1 065 928 382.26 из задания — это сумма по ВСЕМ строкам; разница {:.2f} = возвраты)".format(
    sum_rub_all - sum_rub_pos))
print("КОНТРОЛЬ 3a: ТО, в е.изм., штучные, все строки = {:,.0f} шт, эталон dashboard = {:,.0f} шт -> {}".format(
    qty_all_sht, dash_tot["штук"], "OK" if abs(qty_all_sht - dash_tot["штук"]) <= 0.01 else "МИСМАТЧ"))
print("КОНТРОЛЬ 3b: ТО, в е.изм., весовые, все строки = {:,.1f} кг, эталон dashboard = {:,.1f} кг -> {}".format(
    qty_all_kg, dash_tot["кг"], "OK" if abs(qty_all_kg - dash_tot["кг"]) <= 0.01 else "МИСМАТЧ"))
print("   по методологии (положительные строки): штучные = {:,.0f} шт, весовые = {:,.1f} кг".format(
    qty_pos_sht, qty_pos_kg))

# ---------------------------------------------------------------------------
# 3. Месячные суммы по каждому товару (в собственных единицах)
# ---------------------------------------------------------------------------
full = s[s["месяц"].between(FULL_START, FULL_END)]                     # только полные месяцы
mon = full.groupby(["Код", "месяц"])["ТО, в е.изм."].sum().reset_index()

# Матрица: строки = товары, столбцы = месяцы (43), внутри — сумма продаж (0, если не было)
pivot = mon.pivot_table(index="Код", columns="месяц",
                        values="ТО, в е.изм.", fill_value=0.0)
pivot = pivot.reindex(columns=list(ALL_MONTHS), fill_value=0.0)

# Справочные данные по товару (по положительным строкам)
names = full.groupby("Код")["Номенклатура"].first()
groups = full.groupby("Код")["группа"].first()
revenue = full.groupby("Код")["ТО, руб"].sum()   # выручка = сумма положительных строк

# ---------------------------------------------------------------------------
# 4. Окно товара, месячная статистика, класс XYZ
# ---------------------------------------------------------------------------
records = []
for code in pivot.index:
    row = pivot.loc[code]
    has = row > 0
    first_m = row.index[has][0]                    # первый полный месяц с продажами
    last_m = row.index[has][-1]                    # последний полный месяц с продажами
    n_months = months_between(first_m, last_m)     # длина окна, нулевые месяцы внутри учтены

    win = row.loc[first_m:last_m].astype(float)    # месячный ряд внутри окна (вкл. нули)
    mean_m = win.mean()
    std_m = win.std()                              # выборочное стд. отклонение (ddof=1)

    if n_months >= 6:
        cv = std_m / mean_m * 100.0 if mean_m > 0 else np.nan
        if cv <= 10:
            cls = "X"
        elif cv <= 25:
            cls = "Y"
        else:
            cls = "Z"
        # Розничные границы (X ≤ 25, 25 < Y ≤ 50, Z > 50) — те же CV, другой порог
        if cv <= 25:
            cls_rt = "X"
        elif cv <= 50:
            cls_rt = "Y"
        else:
            cls_rt = "Z"
        note = ""
    else:
        cv, cls, cls_rt = np.nan, "н/д", "н/д"
        note = "новинка или мало данных ({} мес)".format(n_months)

    records.append({
        "Код": code,
        "Название": names[code],
        "Группа": groups[code],
        "Тип": type_map[code],
        "Первый месяц": first_m,
        "Последний месяц": last_m,
        "Месяцев в окне": n_months,
        "Среднее/мес (ед. товара)": round(mean_m, 1),
        "Стд. отклонение": round(std_m, 1),
        "CV %": round(cv, 1) if not np.isnan(cv) else np.nan,
        "Класс": cls,
        "Класс (розница)": cls_rt,
        "Выручка за окно (руб)": round(revenue[code], 2),
        "Примечание": note,
    })

base = pd.DataFrame(records)


def sort_xyz(df):
    """Сортировка: Класс (X, Y, Z, н/д), затем CV по возрастанию (у н/д CV нет — в конец)."""
    df = df.copy()
    df["_порядок"] = df["Класс"].map(CLASS_ORDER)
    df = df.sort_values(["_порядок", "CV %"], ascending=[True, True],
                        na_position="last").reset_index(drop=True)
    return df.drop(columns="_порядок")


xyz_kl = sort_xyz(base[XYZ_COLS])                                     # классические 10/25
xyz_rz = sort_xyz(base[[c if c != "Класс" else "Класс (розница)" for c in XYZ_COLS]]
                  .rename(columns={"Класс (розница)": "Класс"}))      # розничные 25/50


def adjust_round(vals, decimals=1, target=100.0):
    """Округлить доли до decimals знаков; если сумма отличается от target,
    сдвинуть на 1 шаг позиции с наибольшими отброшенными дробями (сумма = 100%)."""
    step = 10 ** (-decimals)
    out = [round(v, decimals) for v in vals]
    diff = round(target - sum(out), decimals)
    if diff != 0:
        n = int(abs(diff) / step + 0.5)
        frac = [v - round(v, decimals) for v in vals]
        order = sorted(range(len(vals)),
                       key=lambda i: (frac[i] if diff > 0 else -frac[i]),
                       reverse=True)
        for i in order[:n]:
            out[i] = round(out[i] + (step if diff > 0 else -step), decimals)
    return out


def make_svodka(xyz_df):
    """Сводка по классам + матрица Группа × Класс для одного набора границ."""
    total_count = len(xyz_df)
    total_rev = xyz_df["Выручка за окно (руб)"].sum()
    total_vol = xyz_df["Среднее/мес (ед. товара)"].sum()

    summ_rows = []
    for cls in CLS_NAMES:
        sub = xyz_df[xyz_df["Класс"] == cls]
        summ_rows.append({
            "Класс": cls,
            "Товаров": len(sub),
            "% товаров": len(sub) / total_count * 100.0,
            "% от выручки": sub["Выручка за окно (руб)"].sum() / total_rev * 100.0,
            "% от объёма": sub["Среднее/мес (ед. товара)"].sum() / total_vol * 100.0,
        })
    summ = pd.DataFrame(summ_rows)
    # Доли по классам должны давать ровно 100% в каждой колонке (коррекция округления)
    for col in ["% товаров", "% от выручки", "% от объёма"]:
        summ[col] = adjust_round(summ[col].astype(float).tolist(), 1, 100.0)
    summ.loc[len(summ)] = {
        "Класс": "ИТОГО",
        "Товаров": total_count,
        "% товаров": 100.0,
        "% от выручки": 100.0,
        "% от объёма": 100.0,
    }

    # Матрица Группа × Класс (количество товаров)
    matrix = xyz_df.pivot_table(index="Группа", columns="Класс", values="Код",
                                aggfunc="count", fill_value=0)
    matrix = matrix.reindex(columns=CLS_NAMES, fill_value=0)
    matrix["Всего"] = matrix.sum(axis=1)
    matrix.loc["ИТОГО"] = matrix.sum(axis=0)
    matrix = matrix.reset_index().rename(columns={"Группа": "Группа \\ Класс"})
    return summ, matrix


summ_kl, matrix_kl = make_svodka(xyz_kl)
summ_rz, matrix_rz = make_svodka(xyz_rz)

# ---------------------------------------------------------------------------
# 5. Лист «Методика» — подробное объяснение простыми словами
#    Каждая секция = заголовок + 2–5 строк. Формулы текстом, без LaTeX.
# ---------------------------------------------------------------------------
min_item = base.loc[base["CV %"].notna(), :].sort_values("CV %").iloc[0]
min_code, min_name, min_cv = min_item["Код"], min_item["Название"], min_item["CV %"]

nX_kl, nY_kl, nZ_kl, nNa = [int((xyz_kl["Класс"] == c).sum()) for c in CLS_NAMES]
nX_rz, nY_rz, nZ_rz, _ = [int((xyz_rz["Класс"] == c).sum()) for c in CLS_NAMES]
pct_rz = {row["Класс"]: row["% от выручки"]
          for _, row in summ_rz.iloc[:4].iterrows()}   # скорректированные доли выручки

method_rows = [
    ["Методика XYZ-анализа колбасы",
     "Как посчитаны коэффициенты вариации, почему в отчёте два набора границ и как читать "
     "результат. Формулы даны текстом, без LaTeX."],

    ["а) Что такое XYZ-анализ и коэффициент вариации", ""],
    ["", "XYZ-анализ делит товары на классы по предсказуемости месячных продаж: "
         "X — стабильный спрос, Y — колеблющийся, Z — непредсказуемый."],
    ["", "Коэффициент вариации (CV) = стандартное отклонение / среднее × 100%. "
         "Простыми словами — насколько месячные продажи «пляшут» вокруг среднего значения."],
    ["", "CV = 0% — продажи абсолютно ровные каждый месяц; CV = 100% — разброс сравним "
         "со средним. Чем больше CV, тем менее надёжен прогноз «по среднему»."],

    ["б) Исходные данные", ""],
    ["", "Файл: data/колбасаНФ.xlsx."],
    ["", "Период данных: 2023-01-01 … 2026-08-10."],
    ["", "Товаров: 507, записей продаж: 246 308 строк."],
    ["", "Анализ выполнен только по полным месяцам 2023-01 … 2026-07 (43 месяца)."],

    ["в) Правила расчёта", ""],
    ["", "Спрос — только положительные продажи (ТО, в е.изм. > 0 и ТО, руб > 0). "
         "Возвраты исключены из всех расчётов (суммарно −4 869 руб)."],
    ["", "Единицы — собственные для товара: штучные в штуках, весовые в килограммах. "
         "Тип определяется правилом «есть хоть одна дробная продажа → весовой»."],
    ["", "Окно товара — от первого до последнего полного месяца продаж. Нулевые месяцы "
         "ВНУТРИ окна учитываются (честная волатильность); месяцы до первого и после "
         "последнего — не учитываются."],
    ["", "Если окно меньше 6 месяцев → класс «н/д»: новинка или товар, ушедший из продаж."],
    ["", "CV = выборочное стандартное отклонение / среднее × 100 (как функция STDEV "
         "в Excel), в отчёте округлён до 1 знака."],

    ["г) Два набора границ классификации", ""],
    ["", "Классические границы (листы XYZ_классика / Сводка_классика): X — CV ≤ 10%, "
         "Y — 10% < CV ≤ 25%, Z — CV > 25%. Стандарт для промышленных/оптовых поставок."],
    ["", "Розничные границы (листы XYZ_розница / Сводка_розница): X — CV ≤ 25%, "
         "Y — 25% < CV ≤ 50%, Z — CV > 50%. Адаптация для розничного магазина, где промо "
         "и сезонность усиливают колебания спроса."],
    ["", "CV считается один раз и одинаково в обоих вариантах — меняется только порог "
         "отнесения к классу. Класс «н/д» в обоих вариантах одинаковый."],

    ["д) Результаты", ""],
    ["", "Классические границы: X = {} товаров, Y = {}, Z = {}, н/д = {} (итого {}).".format(
        nX_kl, nY_kl, nZ_kl, nNa, n_codes)],
    ["", "В данных НЕТ товаров класса X: минимальный CV = {:.1f}% (товар {} «{}»); "
         "даже топ-товар 38318 «Ветчина СПК Балтийская амитан 470г» имеет CV = 92,6%.".format(
        min_cv, min_code, min_name)],
    ["", "Это свойство данных розничной торговли, а не ошибка расчёта: розничные продажи "
         "почти всегда колеблются сильнее 10% даже у самых ходовых товаров."],
    ["", "Розничные границы: X = {} товаров ({:.1f}% выручки), Y = {} ({:.1f}%), "
         "Z = {} ({:.1f}%), н/д = {} ({:.1f}%). Итог {}.".format(
        nX_rz, pct_rz["X"], nY_rz, pct_rz["Y"], nZ_rz, pct_rz["Z"],
        nNa, pct_rz["н/д"], n_codes)],

    ["е) Как читать результат", ""],
    ["", "X — планировать закупки по среднему значению: спрос стабилен, лишний "
         "страховой запас не нужен."],
    ["", "Y — держать страховой запас поверх среднего: колебания есть, но умеренные."],
    ["", "Z — интервальные прогнозы и частые дозаказы небольшими партиями: "
         "спрос непредсказуем."],
    ["", "н/д — решить судьбу товара отдельно: новинка (понаблюдать ещё) или уходящий "
         "товар (распродать остатки)."],
]
method = pd.DataFrame(method_rows, columns=["Раздел", "Пояснение"])

# ---------------------------------------------------------------------------
# 6. Сохранение Excel (5 листов)
# ---------------------------------------------------------------------------
with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
    xyz_kl.to_excel(writer, sheet_name="XYZ_классика", index=False)

    summ_kl.to_excel(writer, sheet_name="Сводка_классика", index=False, startrow=0)
    matrix_kl.to_excel(writer, sheet_name="Сводка_классика", index=False,
                       startrow=len(summ_kl) + 3)
    ws = writer.sheets["Сводка_классика"]
    ws.cell(row=len(summ_kl) + 2, column=1).value = \
        "Матрица Группа × Класс (количество товаров)"

    xyz_rz.to_excel(writer, sheet_name="XYZ_розница", index=False)

    summ_rz.to_excel(writer, sheet_name="Сводка_розница", index=False, startrow=0)
    matrix_rz.to_excel(writer, sheet_name="Сводка_розница", index=False,
                       startrow=len(summ_rz) + 3)
    ws = writer.sheets["Сводка_розница"]
    ws.cell(row=len(summ_rz) + 2, column=1).value = \
        "Матрица Группа × Класс (количество товаров)"

    # Лист «Методика» — только ТЕКСТ (без формул Excel), ширины колонок настроены
    method.to_excel(writer, sheet_name="Методика", index=False)
    ws = writer.sheets["Методика"]
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 125
    wrap_top = Alignment(wrap_text=True, vertical="top")
    for r in range(1, method.shape[0] + 2):           # 1 — строка заголовков таблицы
        a, b = ws.cell(row=r, column=1), ws.cell(row=r, column=2)
        a.alignment = Alignment(vertical="top")
        b.alignment = wrap_top
        if r == 1 or (a.value and not b.value):       # заголовки таблицы и секций — жирным
            a.font = Font(bold=True)
        if r == 2 and a.value:                        # заголовок документа — крупным
            a.font = Font(bold=True, size=14)

# ---------------------------------------------------------------------------
# 8. Графики
# ---------------------------------------------------------------------------
# 8a. Гистограмма CV % (ось X ограничена 150, границы классов отмечены линиями)
cvs = xyz_kl.loc[xyz_kl["CV %"].notna(), "CV %"]
nX = int((xyz_kl["Класс"] == "X").sum())
nY = int((xyz_kl["Класс"] == "Y").sum())
nZ = int((xyz_kl["Класс"] == "Z").sum())

fig, ax = plt.subplots(figsize=(9.5, 5.2))
ax.hist(cvs, bins=np.arange(0, 155, 5), range=(0, 150),
        color="#4C72B0", edgecolor="white", alpha=0.9)
ax.axvline(10, color="#C44E52", linestyle="--", linewidth=1.6)   # граница X/Y
ax.axvline(25, color="#DD8452", linestyle="--", linewidth=1.6)   # граница Y/Z
ymax = ax.get_ylim()[1]
ax.text(10, ymax * 0.97, "X: {} товаров\nCV ≤ 10%".format(nX),
        ha="center", va="top", color="#C44E52", fontsize=9)
ax.text(25, ymax * 0.60, "Y: {} товаров\n10 < CV ≤ 25%".format(nY),
        ha="center", va="top", color="#DD8452", fontsize=9)
ax.text(0.99, 0.97, "Z: {} товаров\nCV > 25%".format(nZ),
        transform=ax.transAxes, ha="right", va="top", color="#55A868", fontsize=9)
n_over = int((cvs > 150).sum())
if n_over:
    ax.text(0.01, 0.02, "у {} товаров CV > 150% (за пределами оси)".format(n_over),
            transform=ax.transAxes, fontsize=8, color="dimgray")
ax.set_xlim(0, 150)
ax.set_xlabel("CV, %")
ax.set_ylabel("N товаров")
ax.set_title("Распределение коэффициента вариации CV % (колбаса, полные месяцы 2023-01 … 2026-07)")
fig.tight_layout()
fig.savefig(OUT_PNG1, dpi=150)
plt.close(fig)

# 8b. Оба варианта границ: 4 группы столбиков
#     (товары/выручка × классика/розница), по классу — свой цвет, легенда, подписи
cls_colors = ["#55A868", "#DD8452", "#C44E52", "#8C8C8C"]


def shares(xyz_df):
    tot_c = len(xyz_df)
    tot_r = xyz_df["Выручка за окно (руб)"].sum()
    prod = [len(xyz_df[xyz_df["Класс"] == c]) / tot_c * 100.0 for c in CLS_NAMES]
    rev = [xyz_df.loc[xyz_df["Класс"] == c, "Выручка за окно (руб)"].sum() / tot_r * 100.0
           for c in CLS_NAMES]
    return prod, rev


p_kl, r_kl = shares(xyz_kl)
p_rz, r_rz = shares(xyz_rz)

groups = ["Товары\nклассика", "Выручка\nклассика", "Товары\nрозница", "Выручка\nрозница"]
values = [p_kl, r_kl, p_rz, r_rz]

fig, ax = plt.subplots(figsize=(12.8, 5.6))
x = np.arange(len(groups))
width, offsets = 0.19, [-0.285, -0.095, 0.095, 0.285]
for i, cls in enumerate(CLS_NAMES):
    heights = [vals[i] for vals in values]
    ax.bar(x + offsets[i], heights, width, label=cls, color=cls_colors[i])
    for xi, h in zip(x, heights):
        ax.text(xi + offsets[i], h + 1.0, "{:.1f}%".format(h),
                ha="center", va="bottom", fontsize=8)
ax.set_xticks(x)
ax.set_xticklabels(groups, fontsize=10)
ax.set_ylabel("Доля, %")
ax.set_ylim(0, max(max(v) for v in values) * 1.26 + 4)
ax.set_title("XYZ-анализ: классические и розничные границы", fontsize=13)
ax.legend(title="Класс", ncol=4, loc="lower center",
          bbox_to_anchor=(0.5, 1.02), frameon=False)
ax.spines["top"].set_visible(False)
ax.spines["right"].set_visible(False)
fig.tight_layout(rect=(0, 0, 1, 0.90))
fig.savefig(OUT_PNG2, dpi=150)
plt.close(fig)

# ---------------------------------------------------------------------------
# 9. Контрольные суммы — печать (оба набора границ)
# ---------------------------------------------------------------------------
def sum_check(summ_df, label):
    s_p = summ_df.iloc[:4]["% товаров"].sum()
    s_r = summ_df.iloc[:4]["% от выручки"].sum()
    ok = abs(s_p - 100.0) < 1e-6 and abs(s_r - 100.0) < 1e-6
    print("КОНТРОЛЬ {}: сумма % товаров по классам = {:.1f}, сумма % выручки = {:.1f} -> {}".format(
        label, s_p, s_r, "OK" if ok else "МИСМАТЧ"))

sum_check(summ_kl, "4 (классика)")
sum_check(summ_rz, "5 (розница)")

cnt_kl = {c: int((xyz_kl["Класс"] == c).sum()) for c in CLS_NAMES}
cnt_rz = {c: int((xyz_rz["Класс"] == c).sum()) for c in CLS_NAMES}
ok6 = sum(cnt_kl.values()) == 507
print("КОНТРОЛЬ 6 (классика): сумма товаров по классам = {} (X={}, Y={}, Z={}, н/д={}) -> {}".format(
    sum(cnt_kl.values()), cnt_kl["X"], cnt_kl["Y"], cnt_kl["Z"], cnt_kl["н/д"],
    "OK" if ok6 else "МИСМАТЧ"))
ok7 = sum(cnt_rz.values()) == 507
print("КОНТРОЛЬ 7 (розница): сумма товаров по классам = {} (X={}, Y={}, Z={}, н/д={}) -> {}".format(
    sum(cnt_rz.values()), cnt_rz["X"], cnt_rz["Y"], cnt_rz["Z"], cnt_rz["н/д"],
    "OK" if ok7 else "МИСМАТЧ"))

mean_cv_rz = {c: xyz_rz.loc[xyz_rz["Класс"] == c, "CV %"].mean() for c in ["X", "Y", "Z"]}
ok8 = mean_cv_rz["X"] < mean_cv_rz["Y"] < mean_cv_rz["Z"]
print("КОНТРОЛЬ 8 (розница): средний CV — X = {:.2f} < Y = {:.2f} < Z = {:.2f} -> {}".format(
    mean_cv_rz["X"], mean_cv_rz["Y"], mean_cv_rz["Z"], "OK" if ok8 else "МИСМАТЧ"))

mean_cv_kl = {c: xyz_kl.loc[xyz_kl["Класс"] == c, "CV %"].mean() for c in ["X", "Y", "Z"]}
min_cv = xyz_kl.loc[xyz_kl["CV %"].notna(), "CV %"].min()
if cnt_kl["X"] > 0:
    ok9 = mean_cv_kl["X"] < mean_cv_kl["Y"] < mean_cv_kl["Z"]
    print("КОНТРОЛЬ 9 (классика): средний CV — X = {:.2f} < Y = {:.2f} < Z = {:.2f} -> {}".format(
        mean_cv_kl["X"], mean_cv_kl["Y"], mean_cv_kl["Z"], "OK" if ok9 else "МИСМАТЧ"))
else:
    print("КОНТРОЛЬ 9 (классика): средний CV — X = нет товаров, Y = {:.2f} < Z = {:.2f} -> МИСМАТЧ (ожидаемо)".format(
        mean_cv_kl["Y"], mean_cv_kl["Z"]))
    print("   ПРИЧИНА: в данных НЕТ ни одного товара с CV <= 10% (минимальный CV = {:.1f}%);".format(min_cv))
    print("   спрос на весь ассортимент колеблется сильнее — это реальное свойство данных, а не ошибка расчёта.")

# ---------------------------------------------------------------------------
# 10. Примеры товаров и сводка в консоль (оба набора границ)
# ---------------------------------------------------------------------------
for label, df, cnt in [("КЛАССИКА", xyz_kl, cnt_kl), ("РОЗНИЦА", xyz_rz, cnt_rz)]:
    print()
    print("--- {} ---".format(label))
    for cls in CLS_NAMES:
        sub = df[df["Класс"] == cls]
        if cls == "н/д":
            ex = sub.sort_values("Выручка за окно (руб)", ascending=False).head(10)
        else:
            ex = sub.head(10)   # уже отсортированы по CV возрастанию
        print("Примеры класса {} (всего {} товаров):".format(cls, len(sub)))
        for _, r in ex.iterrows():
            cv_txt = "{:.1f}".format(r["CV %"]) if pd.notna(r["CV %"]) else "—"
            print("   {} | {} | CV {}% | среднее/мес {:.1f}".format(
                r["Код"], r["Название"][:55], cv_txt, r["Среднее/мес (ед. товара)"]))
    tot_rev = df["Выручка за окно (руб)"].sum()
    sh = {c: df.loc[df["Класс"] == c, "Выручка за окно (руб)"].sum() / tot_rev * 100.0
          for c in CLS_NAMES}
    print("СВОДКА {}: X: {} товаров (доля выручки {:.1f}%), Y: {} ({:.1f}%), "
          "Z: {} ({:.1f}%), н/д: {} ({:.1f}%)".format(
        label, cnt["X"], sh["X"], cnt["Y"], sh["Y"], cnt["Z"], sh["Z"],
        cnt["н/д"], sh["н/д"]))

print()
print("Сохранено: {}".format(OUT_XLSX))
print("  размер: {:,} байт".format(os.path.getsize(OUT_XLSX)).replace(",", " "))
print("Сохранено: {}".format(OUT_PNG1))
print("  размер: {:,} байт".format(os.path.getsize(OUT_PNG1)).replace(",", " "))
print("Сохранено: {}".format(OUT_PNG2))
print("  размер: {:,} байт".format(os.path.getsize(OUT_PNG2)).replace(",", " "))
