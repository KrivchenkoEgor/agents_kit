# -*- coding: utf-8 -*-
"""
29_abc_xyz_matrix.py — матрица ABC × XYZ для колбасы
=====================================================
Совмещает два анализа в одну карту решений:

  ABC — ВАЖНОСТЬ товара (доля в выручке):
      A — первые 80% выручки (звёзды), B — до 95% (середняки),
      C — остальные 5% (аутсайдеры). Считается по сумме «ТО, руб»
      ПОЛОЖИТЕЛЬНЫХ строк за ВЕСЬ период данных.

  XYZ — СТАБИЛЬНОСТЬ спроса (из готового отчёта scripts/28_xyz_analysis.py):
      X — ровный спрос, Y — колеблется, Z — непредсказуемый, н/д — мало данных.
      Классы НЕ пересчитываются: читаются из output/Отчет_XYZ_колбаса.xlsx
      (листы XYZ_классика и XYZ_розница) и сливаются с ABC по коду товара.

Итоговая ячейка = ABC + XYZ (например AX, CZ; при XYZ = н/д — «A н/д»).
Комбинация показывает, КАКОЙ товар по значимости (ABC) и КАК предсказуемо
он продаётся (XYZ) — отсюда и практические рекомендации по закупкам.

Выход:
  output/Отчет_ABC_XYZ_колбаса.xlsx  (листы Матрица, Сводка_классика,
                                      Сводка_розница, Рекомендации, Методика)
  plots/ABC_XYZ_матрица_колбаса.png  (тепловая карта 3×3, классические границы)
"""

import os

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl.styles import Alignment, Font

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data", "колбасаНФ.xlsx")
XYZ_XLSX = os.path.join(ROOT, "output", "Отчет_XYZ_колбаса.xlsx")
OUT_XLSX = os.path.join(ROOT, "output", "Отчет_ABC_XYZ_колбаса.xlsx")
OUT_PNG = os.path.join(ROOT, "plots", "ABC_XYZ_матрица_колбаса.png")

plt.rcParams["font.family"] = "DejaVu Sans"  # шрифт поддерживает кириллицу

# Порядок ячеек матрицы (сортировка листа «Матрица»)
CELL_ORDER = ["AX", "AY", "AZ", "BX", "BY", "BZ",
              "CX", "CY", "CZ", "A н/д", "B н/д", "C н/д"]
ABC_ORDER = ["A", "B", "C"]
XYZ_ORDER = ["X", "Y", "Z", "н/д"]

# ---------------------------------------------------------------------------
# 1. ABC-анализ по сырым данным (положительные строки, весь период)
# ---------------------------------------------------------------------------
df = pd.read_excel(DATA, dtype={"Код ": str})   # «Код » — колонка с пробелом
df["Код"] = df["Код "].str.strip().str.zfill(5)

# СПРОС = только положительные продажи (возвраты исключены, как в XYZ-отчёте)
s = df[(df["ТО, в е.изм."] > 0) & (df["ТО, руб"] > 0)].copy()

# Справочник по товару — ТОЛЬКО из исходника по коду товара:
#   Название/Группа — доминирующее значение по числу строк (value_counts().index[0]);
#   Тип — «весовой», если есть хоть одна дробная продажа «ТО, в е.изм.», иначе
#         «штучный» (формат БЕЗ суффиксов «, шт»/«, кг»);
#   Выручка за ВЕСЬ период — сумма «ТО, руб» положительных строк.
names = df.groupby("Код")["Номенклатура"].apply(
    lambda x: x.value_counts().index[0])
groups = df.groupby("Код")["группа"].apply(
    lambda x: x.value_counts().index[0])
tip = df.groupby("Код")["ТО, в е.изм."].apply(
    lambda x: "весовой" if (x != x.round()).any() else "штучный")
revenue = s.groupby("Код")["ТО, руб"].sum()

# Класс по кумулятивной доле выручки (правило 80/15/5, как в 10_abc_analysis.py):
# A — пока кумулятивная доля <= 80%, B — до 95%, C — остальные.
rev_sorted = revenue.sort_values(ascending=False)
cum = rev_sorted.cumsum() / rev_sorted.sum()
abc = pd.Series("C", index=rev_sorted.index, dtype=object)
abc[cum <= 0.80] = "A"
abc[(cum > 0.80) & (cum <= 0.95)] = "B"

# ЕДИНЫЙ DataFrame с кодом как индексом: все серии выравниваются по общему
# индексу «Код», ручное выравнивание не нужно. (Раньше в dict-конструктор
# передавался pd.Index — он НЕ участвует в выравнивании по индексу, из-за чего
# колонка «Код» и справочные колонки разъезжались.)
abc_df = pd.DataFrame({
    "Название": names,
    "Группа": groups,
    "Тип": tip,
    "Выручка за всё время (руб)": revenue,
    "Класс ABC": abc,
})
abc_df.index.name = "Код"
abc_df = abc_df.reset_index()

# ---------------------------------------------------------------------------
# 2. Чтение готовых XYZ-классов (без пересчёта)
# ---------------------------------------------------------------------------
def read_xyz(sheet):
    """Читает лист XYZ-отчёта: Код (zfill(5)), CV % и Класс (исходные имена)."""
    x = pd.read_excel(XYZ_XLSX, sheet_name=sheet)
    x["Код"] = x["Код"].astype(str).str.strip().str.zfill(5)
    return x[["Код", "CV %", "Класс"]]


xyz_kl = read_xyz("XYZ_классика").rename(
    columns={"CV %": "CV % (XYZ_классика)", "Класс": "Класс (XYZ_классика)"})
xyz_rz = read_xyz("XYZ_розница").rename(
    columns={"CV %": "CV % (XYZ_розница)", "Класс": "Класс (XYZ_розница)"})

# Слияние ABC + XYZ по коду (5-значный код с обеих сторон)
mat = abc_df.merge(xyz_kl, on="Код", how="left").merge(xyz_rz, on="Код", how="left")

# Проверка порядка после merge: merge по коду не должен сдвигать строки
merge_order_ok = bool((mat["Код"].values == abc_df["Код"].values).all())

# Ячейка = ABC + XYZ; XYZ = н/д → «A н/д» и т.п. (пробел только перед н/д)
def make_cell(abc_cls, xyz_cls):
    if xyz_cls == "н/д":
        return "{} н/д".format(abc_cls)
    return abc_cls + xyz_cls


mat["Ячейка (классика)"] = [make_cell(a, z) for a, z in
                            zip(mat["Класс ABC"], mat["Класс (XYZ_классика)"])]
mat["Ячейка (розница)"] = [make_cell(a, z) for a, z in
                           zip(mat["Класс ABC"], mat["Класс (XYZ_розница)"])]

# ---------------------------------------------------------------------------
# 3. Рекомендации — короткая фраза для листа «Матрица»
# ---------------------------------------------------------------------------
REC_SHORT = {
    "AX": "Планировать точно по среднему",
    "AY": "Страховой запас 1,5–2 недели",
    "AZ": "Особый контроль, не допускать дефицита",
    "BX": "Средний приоритет, запас с поправкой",
    "BY": "Средний приоритет, страховой запас",
    "BZ": "Средний приоритет, частые дозаказы",
    "CX": "Рассмотреть снятие, если не нужен для полки",
    "CY": "Кандидат на снятие",
    "CZ": "Кандидат на снятие, минимизировать остатки",
    "A н/д": "Решить судьбу: важно, но мало данных",
    "B н/д": "Решить судьбу: понаблюдать",
    "C н/д": "Решить судьбу: кандидат на распродажу",
}
mat["Рекомендация"] = mat["Ячейка (классика)"].map(REC_SHORT)

# ---------------------------------------------------------------------------
# 4. Сортировка листа «Матрица»: ячейки в порядке AX…C н/д, внутри — выручка ↓
# ---------------------------------------------------------------------------
mat["_порядок"] = mat["Ячейка (классика)"].map(
    {c: i for i, c in enumerate(CELL_ORDER)})
mat = mat.sort_values(["_порядок", "Выручка за всё время (руб)"],
                      ascending=[True, False]).reset_index(drop=True)

matrix_out = mat[[
    "Код", "Название", "Группа", "Тип", "Выручка за всё время (руб)",
    "Класс ABC", "CV % (XYZ_классика)", "Класс (XYZ_классика)",
    "Класс (XYZ_розница)", "Ячейка (классика)", "Ячейка (розница)",
    "Рекомендация",
]].rename(columns={
    "CV % (XYZ_классика)": "CV %",
    "Класс (XYZ_классика)": "Класс XYZ (классика)",
    "Класс (XYZ_розница)": "Класс XYZ (розница)",
})

# ---------------------------------------------------------------------------
# 5. Сводки: 3×4 матрица «N товаров (доля выручки %)» + блок долей выручки
# ---------------------------------------------------------------------------
def adjust_round(vals, decimals=1, target=100.0):
    """Округлить доли до decimals знаков; если сумма отличается от target,
    сдвинуть на 1 шаг позиции с наибольшими отброшенными дробями (сумма = 100%)."""
    step = 10 ** (-decimals)
    out = [round(v, decimals) for v in vals]
    diff = round(target - sum(out), decimals)
    if diff != 0:
        n = int(abs(diff) / step + 0.5)
        frac = [v - round(v, decimals) for v in vals]
        order = sorted(range(len(vals)),
                       key=lambda i: (frac[i] if diff > 0 else -frac[i]),
                       reverse=True)
        for i in order[:n]:
            out[i] = round(out[i] + (step if diff > 0 else -step), decimals)
    return out


def build_svodka(mat_df, cell_col):
    """Сводка ABC × XYZ для одного набора границ.
    Возвращает: (таблица «N (P%)» с итогами, блок долей выручки %,
                 словарь счётчиков по ячейкам, словарь долей выручки по ячейкам)."""
    total_rev = mat_df["Выручка за всё время (руб)"].sum()
    counts = {}
    shares = {}
    for cell in CELL_ORDER:
        sub = mat_df[mat_df[cell_col] == cell]
        counts[cell] = len(sub)
        shares[cell] = sub["Выручка за всё время (руб)"].sum() / total_rev * 100.0

    # Доли по 12 ячейкам должны давать ровно 100.0 (коррекция округления)
    adj = adjust_round([shares[c] for c in CELL_ORDER], 1, 100.0)
    shares_adj = dict(zip(CELL_ORDER, adj))

    # Таблица: строки A/B/C, столбцы X/Y/Z/н/д, каждая ячейка «N (P%)»
    rows = [["", "X", "Y", "Z", "н/д", "Итого"]]
    for a in ABC_ORDER:
        line = [a]
        for z in XYZ_ORDER:
            c = make_cell(a, z)
            line.append("{} ({:.1f}%)".format(counts[c], shares_adj[c]))
        # Итого по строке: сумма товаров и долей ячеек строки
        row_cells = [make_cell(a, z) for z in XYZ_ORDER]
        line.append("{} ({:.1f}%)".format(
            sum(counts[c] for c in row_cells),
            sum(shares_adj[c] for c in row_cells)))
        rows.append(line)
    # Итоговая строка
    total_line = ["Итого"]
    for z in XYZ_ORDER:
        col_cells = [make_cell(a, z) for a in ABC_ORDER]
        total_line.append("{} ({:.1f}%)".format(
            sum(counts[c] for c in col_cells),
            sum(shares_adj[c] for c in col_cells)))
    total_line.append("{} (100.0%)".format(len(mat_df)))
    rows.append(total_line)

    # Отдельный блок: доли выручки по ячейкам матрицы (числа %)
    share_rows = [["Доля выручки %", "X", "Y", "Z", "н/д", "Итого"]]
    for a in ABC_ORDER:
        line = [a]
        for z in XYZ_ORDER:
            line.append(shares_adj[make_cell(a, z)])
        line.append(sum(shares_adj[make_cell(a, z)] for z in XYZ_ORDER))
        share_rows.append(line)
    share_rows.append(["Итого"] + [sum(shares_adj[make_cell(a, z)] for a in ABC_ORDER)
                                   for z in XYZ_ORDER] + [100.0])

    return rows, share_rows, counts, shares_adj


svod_kl_rows, share_kl_rows, cnt_kl, shr_kl = build_svodka(mat, "Ячейка (классика)")
svod_rz_rows, share_rz_rows, cnt_rz, shr_rz = build_svodka(mat, "Ячейка (розница)")

# ---------------------------------------------------------------------------
# 6. Лист «Рекомендации»: 9+3 строк — по одной на каждую ячейку (классика)
# ---------------------------------------------------------------------------
REC_LONG = {
    "AX": "Самый ценный и стабильный товар: планируйте закупки точно по среднему "
          "месячному спросу, страховой запас минимальный. Автозаказ отлично "
          "справляется, ручное вмешательство не требуется.",
    "AY": "Важный товар с умеренными колебаниями: держите страховой запас на "
          "1,5–2 недели продаж сверх среднего. Следите за оборачиваемостью — "
          "дефицит здесь дороже излишка.",
    "AZ": "Топ-товар с непредсказуемым спросом: используйте интервальный прогноз "
          "(нижняя граница как точка дозаказа), заказывайте чаще небольшими "
          "партиями. Не допускайте дефицита — потерянная выручка перекрывает "
          "любые затраты на страховой запас.",
    "BX": "Стабильный товар средней важности: запас с небольшой поправкой на "
          "колебания. Проверяйте регулярно, но детальный контроль не нужен.",
    "BY": "Средний приоритет при умеренных колебаниях: страховой запас и "
          "периодический пересмотр объёма закупки. Оптимизируйте партии по "
          "фактической оборачиваемости.",
    "BZ": "Средний приоритет при непредсказуемом спросе: запас с поправкой на "
          "волатильность, частые дозаказы небольшими партиями. Избегайте "
          "крупных единоразовых закупок.",
    "CX": "Стабильный, но малозначимый товар: если он не нужен для полноты "
          "полки и не входит в постоянный ассортимент — рассмотрите снятие. "
          "Затраты на содержание превышают выгоду.",
    "CY": "Низкая выручка при колеблющемся спросе: кандидат на снятие. "
          "Сократите закуп до минимальных партий и пересмотрите позицию в "
          "ассортименте.",
    "CZ": "Низкая выручка при непредсказуемом спросе: кандидат на снятие. "
          "Минимизируйте остатки, распродавайте текущий запас, не пополняйте.",
    "A н/д": "Товар с большой выручкой, но мало данных (новинка или уходит): "
             "выясните причину — если новинка, дайте ему время и понаблюдайте "
             "ещё 2–3 месяца; если уходит — спланируйте распродажу остатков "
             "без ущерба для выручки.",
    "B н/д": "Средняя выручка при малом количестве данных: понаблюдайте ещё "
             "несколько месяцев. Решите судьбу: развивать (увеличить закуп) "
             "или постепенно выводить.",
    "C н/д": "Низкая выручка и мало данных: решите судьбу сразу. Если товар "
             "не нужен для полноты ассортимента — распродайте остатки и "
             "снимите с закупки.",
}

rec_rows = [["Ячейка", "Товаров", "Доля выручки %", "Рекомендация"]]
for cell in CELL_ORDER:
    rec_rows.append([cell, cnt_kl[cell], shr_kl[cell], REC_LONG[cell]])

# ---------------------------------------------------------------------------
# 7. Лист «Методика» — краткое описание (5–8 строк)
# ---------------------------------------------------------------------------
method_rows = [
    ["Что такое ABC", "ABC — важность товара: A — первые 80% выручки, "
     "B — до 95%, C — остальные 5% (пулы 80/15/5)."],
    ["Что такое XYZ", "XYZ — стабильность спроса: X — ровные продажи, "
     "Y — колеблются, Z — непредсказуемые. Класс н/д — окно меньше 6 месяцев."],
    ["Период ABC", "Весь период данных: 2023-01 … 2026-07. Выручка = сумма "
     "ТО, руб только по положительным строкам (возвраты исключены)."],
    ["Период XYZ", "Окно товара от первого до последнего полного месяца продаж; "
     "CV = стд. отклонение / среднее × 100; нулевые месяцы внутри окна учтены."],
    ["Границы классика", "X — CV ≤ 10%, Y — 10% < CV ≤ 25%, Z — CV > 25% "
     "(стандарт для оптовых поставок)."],
    ["Границы розница", "X — CV ≤ 25%, Y — 25% < CV ≤ 50%, Z — CV > 50% "
     "(адаптация для розничного магазина)."],
    ["Ячейка матрицы", "ABC + XYZ, например AX (важный и стабильный) или "
     "CZ (незначимый и непредсказуемый). XYZ = н/д → «A н/д» и т.п."],
    ["Полный XYZ-отчёт", "output/Отчет_XYZ_колбаса.xlsx (листы XYZ_классика, "
     "XYZ_розница, Сводки, Методика) — там все CV и классы по каждому товару."],
]

# ---------------------------------------------------------------------------
# 8. Сохранение Excel (5 листов)
# ---------------------------------------------------------------------------
with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
    matrix_out.to_excel(writer, sheet_name="Матрица", index=False)
    ws = writer.sheets["Матрица"]
    ws.freeze_panes = "A2"
    widths = {"A": 8, "B": 46, "C": 20, "D": 11, "E": 20,
              "F": 10, "G": 8, "H": 17, "I": 17, "J": 13, "K": 13, "L": 40}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for c in range(1, len(matrix_out.columns) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    def write_block(writer, sheet_name, rows):
        """Записать 2D-список строк на лист (создаёт лист, если его нет).
        Возвращает рабочий лист."""
        if sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
        else:
            ws = writer.book.create_sheet(sheet_name)
        for r, row in enumerate(rows, start=1):
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                if r == 1:
                    cell.font = Font(bold=True)
                if isinstance(val, float):
                    cell.number_format = "0.0"
        for r in range(1, len(rows) + 1):
            ws.cell(row=r, column=1).font = Font(bold=True)
        ws.column_dimensions["A"].width = 12
        for c in range(2, len(rows[0]) + 1):
            ws.column_dimensions[chr(64 + c)].width = 18
        return ws

    def write_share_block(ws, share_rows, start_row):
        """Доли выручки по ячейкам — отдельным блоком под таблицей сводки."""
        for r, row in enumerate(share_rows, start=start_row):
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                if r == start_row:
                    cell.font = Font(bold=True)
                if isinstance(val, float):
                    cell.number_format = "0.0"
            ws.cell(row=r, column=1).font = Font(bold=True)

    ws_kl = write_block(writer, "Сводка_классика", svod_kl_rows)
    write_share_block(ws_kl, share_kl_rows, len(svod_kl_rows) + 3)

    ws_rz = write_block(writer, "Сводка_розница", svod_rz_rows)
    write_share_block(ws_rz, share_rz_rows, len(svod_rz_rows) + 3)

    write_block(writer, "Рекомендации", rec_rows)
    ws = writer.sheets["Рекомендации"]
    for c, w in {"A": 12, "B": 10, "C": 16, "D": 110}.items():
        ws.column_dimensions[c].width = w
    for r in range(2, len(rec_rows) + 1):
        ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")

    write_block(writer, "Методика", method_rows)
    ws = writer.sheets["Методика"]
    for c, w in {"A": 22, "B": 120}.items():
        ws.column_dimensions[c].width = w
    for r in range(2, len(method_rows) + 1):
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")

# ---------------------------------------------------------------------------
# 9. Тепловая карта 3×3 (классические границы), цвет = доля выручки
# ---------------------------------------------------------------------------
grid_counts = np.array([[cnt_kl[a + z] for z in XYZ_ORDER[:3]] for a in ABC_ORDER])
grid_shares = np.array([[shr_kl[a + z] for z in XYZ_ORDER[:3]] for a in ABC_ORDER])

fig, ax = plt.subplots(figsize=(7.5, 6.0))
im = ax.imshow(grid_shares, cmap="YlOrRd", aspect="auto", vmin=0, vmax=100)
for i in range(3):
    for j in range(3):
        ax.text(j, i, "{}\n({:.1f}%)".format(int(grid_counts[i, j]),
                                             grid_shares[i, j]),
                ha="center", va="center", fontsize=13, fontweight="bold",
                color="black" if grid_shares[i, j] < 55 else "white")
ax.set_xticks(range(3))
ax.set_xticklabels(["X", "Y", "Z"])
ax.set_yticks(range(3))
ax.set_yticklabels(["A", "B", "C"])
ax.set_xlabel("XYZ — стабильность спроса")
ax.set_ylabel("ABC — важность (доля выручки)")
ax.set_title("Матрица ABC × XYZ (колбаса, классические границы)\n"
             "N товаров и доля выручки по ячейкам")
cbar = fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
cbar.set_label("доля выручки %")
fig.tight_layout()
fig.savefig(OUT_PNG, dpi=150)
plt.close(fig)

# ---------------------------------------------------------------------------
# 10. Контрольные суммы
# ---------------------------------------------------------------------------
# 1. Всего строк и суммы сводок (товары и доли выручки)
n_rows = len(matrix_out)
ok1_kl = sum(cnt_kl[c] for c in CELL_ORDER) == 507 and \
         abs(sum(shr_kl[c] for c in CELL_ORDER) - 100.0) < 1e-6
ok1_rz = sum(cnt_rz[c] for c in CELL_ORDER) == 507 and \
         abs(sum(shr_rz[c] for c in CELL_ORDER) - 100.0) < 1e-6
print("КОНТРОЛЬ 1a: строк в «Матрица» = {}, ожидание 507 -> {}".format(
    n_rows, "OK" if n_rows == 507 else "МИСМАТЧ"))
print("КОНТРОЛЬ 1b: сводка классика — товаров {} (507), доли выручки {:.1f}% (100.0) -> {}".format(
    sum(cnt_kl[c] for c in CELL_ORDER), sum(shr_kl[c] for c in CELL_ORDER),
    "OK" if ok1_kl else "МИСМАТЧ"))
print("КОНТРОЛЬ 1c: сводка розница  — товаров {} (507), доли выручки {:.1f}% (100.0) -> {}".format(
    sum(cnt_rz[c] for c in CELL_ORDER), sum(shr_rz[c] for c in CELL_ORDER),
    "OK" if ok1_rz else "МИСМАТЧ"))

# 2. Выручка ABC и кумулятивные пороги
tot_rev = revenue.sum()
share_a = revenue[abc == "A"].sum() / tot_rev * 100.0
share_ab = revenue[abc.isin(["A", "B"])].sum() / tot_rev * 100.0
ok2 = abs(tot_rev - 1_065_933_251.43) <= 0.01 and 79 <= share_a <= 81 \
      and 94 <= share_ab <= 96
print("КОНТРОЛЬ 2: выручка ABC (положит. строки) = {:.2f} (эталон 1065933251.43), "
      "доля A = {:.2f}% (79–81), A+B = {:.2f}% (94–96) -> {}".format(
    tot_rev, share_a, share_ab, "OK" if ok2 else "МИСМАТЧ"))

# 3. Попарное сравнение XYZ-классов и CV с отчётом
mism = 0
for sheet, cls_col, cv_col in [("XYZ_классика", "Класс (XYZ_классика)", "CV % (XYZ_классика)"),
                               ("XYZ_розница", "Класс (XYZ_розница)", "CV % (XYZ_розница)")]:
    ref = read_xyz(sheet).rename(columns={"Класс": "k", "CV %": "c"})
    m = mat.merge(ref, on="Код", how="left")
    for _, r in m.iterrows():
        c1, c2 = r[cls_col], r["k"]
        if c1 != c2:
            mism += 1
        v1, v2 = r[cv_col], r["c"]
        if (pd.isna(v1) != pd.isna(v2)) or (not pd.isna(v1) and abs(v1 - v2) > 0.001):
            mism += 1
print("КОНТРОЛЬ 3: расхождений XYZ-классов/CV с отчётом (классика+розница) = {} -> {}".format(
    mism, "OK" if mism == 0 else "МИСМАТЧ"))

# 4. Количество ячеек «н/д» (классика)
n_nd = int((mat["Класс (XYZ_классика)"] == "н/д").sum())
print("КОНТРОЛЬ 4: ячеек н/д (классика) = {}, ожидание 117 -> {}".format(
    n_nd, "OK" if n_nd == 117 else "МИСМАТЧ"))

# 5. Суммы по ячейкам равны количеству товаров ABC-классов.
#    Часть товаров имеет XYZ = «н/д» (всего 117: A=18, B=23, C=76), поэтому
#    корректный инвариант — сумма по ВСЕМ ячейкам класса, включая «X н/д».
sum_a_xyz = sum(cnt_kl[c] for c in ["AX", "AY", "AZ"])
sum_b_xyz = sum(cnt_kl[c] for c in ["BX", "BY", "BZ"])
sum_c_xyz = sum(cnt_kl[c] for c in ["CX", "CY", "CZ"])
n_a = int((mat["Класс ABC"] == "A").sum())
n_b = int((mat["Класс ABC"] == "B").sum())
n_c = int((mat["Класс ABC"] == "C").sum())
sum_a = sum_a_xyz + cnt_kl["A н/д"]
sum_b = sum_b_xyz + cnt_kl["B н/д"]
sum_c = sum_c_xyz + cnt_kl["C н/д"]
ok5 = sum_a == n_a and sum_b == n_b and sum_c == n_c
print("КОНТРОЛЬ 5: ячейки класса A (AX+AY+AZ={}+н/д={}) = {} = A ({}) ; "
      "B (BX+BY+BZ={}+н/д={}) = {} = B ({}) ; C (CX+CY+CZ={}+н/д={}) = {} = C ({}) -> {}".format(
    sum_a_xyz, cnt_kl["A н/д"], sum_a, n_a,
    sum_b_xyz, cnt_kl["B н/д"], sum_b, n_b,
    sum_c_xyz, cnt_kl["C н/д"], sum_c, n_c, "OK" if ok5 else "МИСМАТЧ"))

# 6. Сверка справочных колонок с XYZ-отчётом по коду (Название/Группа/Тип/Класс/CV %)
xyz_ref = pd.read_excel(XYZ_XLSX, sheet_name="XYZ_классика")
xyz_ref["Код"] = xyz_ref["Код"].astype(str).str.strip().str.zfill(5)
xyz_ref = xyz_ref.set_index("Код")
# Тип в XYZ-отчёте с суффиксом («штучный, шт»/«весовой, кг») — нормализуем для сравнения
xyz_ref["Тип_норм"] = xyz_ref["Тип"].str.replace(", шт", "").str.replace(", кг", "")

mm = matrix_out.set_index("Код")
mm = mm.join(xyz_ref[["Название", "Группа", "Тип_норм", "Класс", "CV %"]],
             how="left", lsuffix="_м", rsuffix="_э")
mism_name = int((mm["Название_м"] != mm["Название_э"]).sum())
mism_grp = int((mm["Группа_м"] != mm["Группа_э"]).sum())
mism_tip = int((mm["Тип"] != mm["Тип_норм"]).sum())
mism_cls = int((mm["Класс XYZ (классика)"] != mm["Класс"]).sum())
cv_m = pd.to_numeric(mm["CV %_м"], errors="coerce")
cv_e = pd.to_numeric(mm["CV %_э"], errors="coerce")
mism_cv = int(((cv_m.notna() & cv_e.notna()) & (cv_m - cv_e).abs() > 0.05).sum())
ok6 = (mism_name + mism_grp + mism_tip + mism_cls + mism_cv) == 0
print("КОНТРОЛЬ 6: сверка с XYZ_классика по коду -> Название {} (0), Группа {} (0), "
      "Тип {} (0), Класс {} (0), CV% {} (0, допуск 0.05) -> {}".format(
    mism_name, mism_grp, mism_tip, mism_cls, mism_cv, "OK" if ok6 else "МИСМАТЧ"))

# 7. Сверка контрольных кодов с исходником: название в матрице == единственному в исходнике
CHECK_CODES = ["06193", "38331", "38318", "198909", "176032"]
src_name = df.groupby("Код")["Номенклатура"].apply(lambda x: x.value_counts().index[0])
bad = [c for c in CHECK_CODES
       if matrix_out.set_index("Код").loc[c, "Название"] != src_name[c]]
print("КОНТРОЛЬ 7: сверка 5 кодов с исходником (06193, 38331, 38318, 198909, 176032) "
      "-> расхождений {} -> {}".format(len(bad), "OK" if not bad else "МИСМАТЧ: " + ", ".join(bad)))

# 8. Порядок строк после merge с XYZ не сдвинут
print("КОНТРОЛЬ 8: порядок строк после merge с XYZ не сдвинут -> {}".format(
    "OK" if merge_order_ok else "МИСМАТЧ"))

# ---------------------------------------------------------------------------
# 11. Печать сводки классика и размеров файлов
# ---------------------------------------------------------------------------
print()
print("СВОДКА_КЛАССИКА (N товаров, доля выручки):")
for row in svod_kl_rows:
    print("  " + " | ".join("{:<13}".format(str(v)) for v in row))
print()
print("СВОДКА_РОЗНИЦА (N товаров, доля выручки):")
for row in svod_rz_rows:
    print("  " + " | ".join("{:<13}".format(str(v)) for v in row))

print()
print("Сохранено: {}".format(OUT_XLSX))
print("  размер: {:,} байт".format(os.path.getsize(OUT_XLSX)).replace(",", " "))
print("Сохранено: {}".format(OUT_PNG))
print("  размер: {:,} байт".format(os.path.getsize(OUT_PNG)).replace(",", " "))
